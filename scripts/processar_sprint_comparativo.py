import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

print("======================================")
print("CONSOLIDACAO DE SPRINT DOS PROJETOS")
print("======================================")

# ============================================================
# DEFINIR PASTA BASE
# ============================================================

base_dir = os.path.join(os.getcwd(), "saida")

if not os.path.exists(base_dir):
    print(f"❌ Pasta não encontrada: {base_dir}")
    input("Pressione ENTER...")
    exit()

# ============================================================
# LISTAR PROJETOS
# ============================================================

projetos = [
    p for p in os.listdir(base_dir)
    if os.path.isdir(os.path.join(base_dir, p))
    and p != "Consolidado_Sprint"
]

df_final_total = pd.DataFrame()

# ============================================================
# LOOP EM TODOS OS PROJETOS
# ============================================================

for projeto in projetos:

    print(f"\n🔄 Processando projeto: {projeto}")

    caminho_projeto = os.path.join(base_dir, projeto)

    arquivos = [
        f for f in os.listdir(caminho_projeto)
        if f.endswith(".xlsx")
        and "Sprint_" not in f
    ]

    if not arquivos:
        print("⚠ Nenhum arquivo encontrado, pulando...")
        continue

    arquivos_completos = [
        os.path.join(caminho_projeto, f) for f in arquivos
    ]

    arquivo_recente = max(arquivos_completos, key=os.path.getmtime)

    print(f"📄 Arquivo usado: {os.path.basename(arquivo_recente)}")

    # ============================================================
    # LER SOMENTE AS ABAS CORRETAS
    # ============================================================

    try:
        df_atual = pd.read_excel(
            arquivo_recente,
            sheet_name="Sprint Atual",
            engine="openpyxl"
        )

        df_proximo = pd.read_excel(
            arquivo_recente,
            sheet_name="Próximo Sprint",
            engine="openpyxl"
        )

    except Exception as e:
        print(f"❌ Erro ao ler abas: {e}")
        continue

    # ============================================================
    # FILTRO EMISSÃO INICIAL
    # ============================================================

    df_atual = df_atual.dropna(subset=["Nome da Tarefa"])
    df_proximo = df_proximo.dropna(subset=["Nome da Tarefa"])

    df_atual = df_atual[
        df_atual["Nome da Tarefa"].str.lower() == "emissão inicial"
    ]

    df_proximo = df_proximo[
        df_proximo["Nome da Tarefa"].str.lower() == "emissão inicial"
    ]

    # ============================================================
    # DEFINIR SPRINT
    # ============================================================

    df_atual["Sprint"] = "Sprint Atual"
    df_proximo["Sprint"] = "Próximo Sprint"

    # ============================================================
    # UNIR DADOS
    # ============================================================

    df = pd.concat([df_atual, df_proximo], ignore_index=True)

    # ============================================================
    # ESTRUTURAR PADRÃO FINAL
    # ============================================================

    df_final = pd.DataFrame({
        "Projeto": projeto,
        "Documento": df["Nome do Resumo da Tarefa"],
        "Disciplina": df["Disciplina"],
        "Etapa": "Emissão inicial",
        "Status": df["Status"],
        "Sprint": df["Sprint"]
    })

    df_final_total = pd.concat([df_final_total, df_final], ignore_index=True)

# ============================================================
# SALVAR CONSOLIDADO
# ============================================================

saida_master = os.path.join(base_dir, "Consolidado_Sprint")

if not os.path.exists(saida_master):
    os.makedirs(saida_master)

data_hoje = datetime.now().strftime("%d-%m-%Y")

arquivo_saida = os.path.join(
    saida_master,
    f"Consolidado_Sprint_{data_hoje}.xlsx"
)

# Salvar Excel
df_final_total.to_excel(arquivo_saida, index=False)

# ============================================================
# CRIAR TABELA AUTOMÁTICA (tbl_sprint)
# ============================================================

wb = load_workbook(arquivo_saida)
ws = wb.active

max_row = ws.max_row
max_col = ws.max_column

end_col = get_column_letter(max_col)

tabela_ref = f"A1:{end_col}{max_row}"

tabela = Table(displayName="tbl_sprint", ref=tabela_ref)

style = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)

tabela.tableStyleInfo = style
ws.add_table(tabela)

wb.save(arquivo_saida)

# ============================================================
# FINALIZAÇÃO
# ============================================================

print("\n✅ PROCESSAMENTO FINALIZADO COM SUCESSO!")
print(f"\n📁 Arquivo gerado:")
print(arquivo_saida)

input("\nPressione ENTER para finalizar...")