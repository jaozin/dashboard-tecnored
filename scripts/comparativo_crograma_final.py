import pandas as pd
import datetime
import os
import numpy as np
import glob
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.axis import DateAxis

# ============================================================
# 1. Seleção automática de projetos (DINÂMICO)
# ============================================================

base_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))

processado_base = os.path.join(base_dir, "Comparativo_Cronograma", "processado")

# Lista automaticamente os projetos (pastas)
projetos = [
    nome for nome in os.listdir(processado_base)
    if os.path.isdir(os.path.join(processado_base, nome))
]

if not projetos:
    raise ValueError("Nenhum projeto encontrado na pasta processado.")

# Ordenar (opcional, mas deixa bonito)
projetos.sort()

print("Escolha o projeto:")

for i, proj in enumerate(projetos, start=1):
    print(f"{i} - {proj}")

opcao = input("Digite o número do projeto: ").strip()

try:
    projeto_nome = projetos[int(opcao) - 1]
except:
    raise ValueError("Projeto inválido!")

# BASE
arquivo1 = os.path.join(base_dir, "Comparativo_Cronograma", "base", projeto_nome, f"AC_{projeto_nome}_Base.xlsx")

if not os.path.exists(arquivo1):
    raise FileNotFoundError(f"Base não encontrada: {arquivo1}")

# PROCESSADO (pega o mais recente)
processado_path = os.path.join(base_dir, "Comparativo_Cronograma", "processado", projeto_nome)

arquivos = glob.glob(os.path.join(processado_path, "*.xlsx"))

if not arquivos:
    raise FileNotFoundError(f"Nenhum arquivo encontrado em {processado_path}")

def extrair_data(nome_arquivo):
    # Ex: AC_J0006_10-06-2026.xlsx
    base = os.path.basename(nome_arquivo)
    data_str = base.split("_")[-1].replace(".xlsx", "")
    return datetime.datetime.strptime(data_str, "%d-%m-%Y")

arquivo2 = max(arquivos, key=extrair_data)

print(f"\nBase selecionada: {arquivo1}")
print(f"Arquivo atualizado selecionado: {arquivo2}")

# DATA (mantém igual)
print("\nInforme a DATA DE REFERÊNCIA (dd/mm/aaaa):")
data_ref_input = input("Data: ").strip()
data_ref = pd.to_datetime(data_ref_input, dayfirst=True)

# SAÍDA (ajustada)
caminho_base = os.path.join(base_dir, "Comparativo_Cronograma", "saida", projeto_nome)
os.makedirs(caminho_base, exist_ok=True)

# ============================================================
# 2. Ler arquivos e Identificar incluídas/excluídas
# ============================================================
df1 = pd.read_excel(arquivo1)
df2 = pd.read_excel(arquivo2)

col_nome = "Nome da Tarefa"
df1[col_nome] = df1[col_nome].astype(str).str.strip()
df2[col_nome] = df2[col_nome].astype(str).str.strip()

set1, set2 = set(df1[col_nome]), set(df2[col_nome])
incluidas, excluidas = set2 - set1, set1 - set2

df2["Comparação"] = ""
df2.loc[df2[col_nome].isin(incluidas), "Comparação"] = "Tarefa Incluída"
df_excl = df1[df1[col_nome].isin(excluidas)].copy()
df_excl["Comparação"] = "Tarefa Excluída"
df_final = pd.concat([df2, df_excl], ignore_index=True)

# ============================================================
# 3. Limpeza de Datas e Funções Auxiliares
# ============================================================
def limpar_data(valor):
    if pd.isna(valor) or str(valor).strip() == "" or str(valor).lower() == "nan": return pd.NaT
    s = str(valor).split(' ')[-1]
    return pd.to_datetime(s, dayfirst=True, format="%d/%m/%y", errors='coerce')

colunas_datas = ["Início", "Término", "Término da linha de base", "Início_LB_R0", "Término_LB_R0", "Término real"]
for col in colunas_datas:
    if col in df_final.columns:
        df_final[col] = df_final[col].apply(limpar_data)

def valor_pct(x):
    if pd.isna(x): return 0.0
    s = str(x).replace("%", "").replace(",", ".").strip()
    try:
        v = float(s)
        return v if v <= 1.0 else v / 100.0
    except: return 0.0

df_final["pct_calc"] = df_final["% concluída"].apply(valor_pct)

# ============================================================
# 4. Colunas Calculadas
# ============================================================
df_final[col_nome] = df_final[col_nome].astype(str).str.strip().str.lower()
df_final["Status"] = df_final["Status"].astype(str).str.strip()

df_final["pct_val"] = pd.to_numeric(df_final["% concluída"].astype(str).str.replace("%", "").str.replace(",", "."), errors='coerce')
df_final["pct_val"] = df_final["pct_val"].apply(lambda x: x / 100 if x > 1 else x)

# Sinalizadores de Emissão
df_final["Emissão inicial total"] = (df_final[col_nome] == "emissão inicial").astype(int)

is_concluida = (df_final["pct_val"] >= 0.99)
is_atrasada = (df_final["Status"] == "Atrasada")

df_final["Emissão inicial concluída"] = ((df_final[col_nome] == "emissão inicial") & is_concluida).astype(int)
df_final["Emissão Inicial Atrasada"] = ((df_final[col_nome] == "emissão inicial") & is_atrasada).astype(int)

df_final["Aprovação total"] = (df_final[col_nome] == "aprovação").astype(int)

mask_avaliacao = (
    (df_final[col_nome] == "avaliação do cliente") |
    (df_final[col_nome] == "análise tecnored")
)
df_final["Avaliação do cliente concluída"] = (mask_avaliacao & is_concluida).astype(int)
df_final["Avaliação Cliente Atrasada"] = (mask_avaliacao & is_atrasada).astype(int)

mask_atendimento = (
    (df_final[col_nome] == "atendimento de comentário") |
    (df_final[col_nome] == "atendimento de comentários")
)
df_final["Atendimento de comentário concluído"] = (mask_atendimento & is_concluida).astype(int)
df_final["Atendimento Atrasado"] = (mask_atendimento & is_atrasada).astype(int)

df_final["Aprovação concluída"] = ((df_final[col_nome] == "aprovação") & is_concluida).astype(int)
df_final["Aprovação Atrasada"] = ((df_final[col_nome] == "aprovação") & is_atrasada).astype(int)

df_final["Aderência"] = (
    (df_final["Emissão inicial concluída"] == 1) & 
    (df_final["Emissão inicial total"] == 1) & 
    (df_final["Término_LB_R0"] <= data_ref)
).astype(int)

mask_previsto = (df_final["Término_LB_R0"] <= data_ref) & (df_final["Emissão inicial total"] == 1)
total_previsto_aderencia = df_final.loc[mask_previsto, "Emissão inicial total"].sum()
total_aderencia = df_final["Aderência"].sum()

aderencia_pct = (total_aderencia / total_previsto_aderencia) if total_previsto_aderencia > 0 else 0
print(f"Cálculo de Aderência concluído: {aderencia_pct:.2%}")

# ============================================================
# 5. Tabelas Resumo, Sprint, Concluídas e Atrasadas
# ============================================================
num_documentos = df_final["Emissão inicial total"].sum()

cols_concl = ["Emissão inicial total", "Emissão inicial concluída", "Avaliação do cliente concluída", "Atendimento de comentário concluído", "Aprovação concluída"]
cols_atras = ["Emissão Inicial Atrasada", "Avaliação Cliente Atrasada", "Atendimento Atrasado", "Aprovação Atrasada"]

df_resumo_concl = df_final.groupby("Disciplina")[cols_concl].sum().reset_index()

# Cálculo das taxas de avanço solicitadas
df_resumo_concl["%EI"] = df_resumo_concl["Emissão inicial concluída"].div(df_resumo_concl["Emissão inicial total"]).fillna(0)
df_resumo_concl["%AP"] = df_resumo_concl["Aprovação concluída"].div(df_resumo_concl["Emissão inicial total"]).fillna(0)

# Alinhamento exato das colunas no Excel
cols_layout_concl = [
    "Disciplina", "Emissão inicial total", "Emissão inicial concluída", 
    "%EI", "Avaliação do cliente concluída", "Atendimento de comentário concluído", 
    "Aprovação concluída", "%AP"
]
df_resumo_concl = df_resumo_concl[cols_layout_concl]

df_resumo_atras = df_final.groupby("Disciplina")[cols_atras].sum().reset_index()

domingo_sub = data_ref + datetime.timedelta(days=(6 - data_ref.weekday() if data_ref.weekday() != 6 else 7))
domingo_seg = domingo_sub + datetime.timedelta(days=7)
cols_sprint_layout = ["Sprint", "Id", "EDT", "Nome do Resumo da Tarefa", "Nome da Tarefa", "Status", "% concluída", "Nomes dos recursos", "Disciplina", "Crítica"]

df_sprint_atual = df_final[(df_final["pct_calc"] < 1) & (df_final["Início"] <= domingo_sub)][cols_sprint_layout]
df_prox_sprint = df_final[(df_final["pct_calc"] < 1) & (df_final["Início"] > domingo_sub) & (df_final["Início"] <= domingo_seg)][cols_sprint_layout]

df_concluidas_list = df_final[df_final["pct_calc"] == 1].copy()
df_atrasadas_list = df_final[(df_final["pct_calc"] < 1) & (df_final["Término"] <= data_ref)].copy()

# ============================================================
# 6. Aba Curva S
# ============================================================
start_date = pd.Timestamp("2026-03-20")
end_date = df_final["Término_LB_R0"].max()

if pd.isna(end_date):
    end_date = df_final["Término da linha de base"].max()

all_dates = pd.date_range(start=start_date, end=end_date, freq='D')
df_curva = pd.DataFrame({"Data": all_dates})

df_prev_emissao = df_final[df_final["Emissão inicial total"] == 1].groupby("Término_LB_R0").size().reset_index(name="Emissão prevista na LB")
df_prev_emissao.rename(columns={"Término_LB_R0": "Data"}, inplace=True)

df_real_emissao = df_final[df_final["Emissão inicial concluída"] == 1].groupby("Término").size().reset_index(name="Emissão realizada por dia")
df_real_emissao.rename(columns={"Término": "Data"}, inplace=True)

df_prev_aprov = df_final[df_final["Aprovação total"] == 1].groupby("Término_LB_R0").size().reset_index(name="Aprovação prevista na LB")
df_prev_aprov.rename(columns={"Término_LB_R0": "Data"}, inplace=True)

df_real_aprov = df_final[df_final["Aprovação concluída"] == 1].groupby("Término").size().reset_index(name="Aprovação realizada por dia")
df_real_aprov.rename(columns={"Término": "Data"}, inplace=True)

df_curva["Data"] = pd.to_datetime(df_curva["Data"]).dt.normalize()
df_prev_emissao["Data"] = pd.to_datetime(df_prev_emissao["Data"]).dt.normalize()
df_real_emissao["Data"] = pd.to_datetime(df_real_emissao["Data"]).dt.normalize()
df_prev_aprov["Data"] = pd.to_datetime(df_prev_aprov["Data"]).dt.normalize()
df_real_aprov["Data"] = pd.to_datetime(df_real_aprov["Data"]).dt.normalize()

df_curva = df_curva.merge(df_prev_emissao, on="Data", how="left").fillna(0)
df_curva = df_curva.merge(df_real_emissao, on="Data", how="left").fillna(0)
df_curva = df_curva.merge(df_prev_aprov, on="Data", how="left").fillna(0)
df_curva = df_curva.merge(df_real_aprov, on="Data", how="left").fillna(0)

df_curva["Emissão prevista na LB acumulada"] = df_curva["Emissão prevista na LB"].cumsum()
mask_ref = (df_curva["Data"] <= data_ref).astype(int)
df_curva["Emissão realizada acumulada"] = (df_curva["Emissão realizada por dia"] * mask_ref).cumsum()

df_curva["Aprovação prevista na LB acumulada"] = df_curva["Aprovação prevista na LB"].cumsum()
df_curva["Aprovação realizada acumulada"] = (df_curva["Aprovação realizada por dia"] * mask_ref).cumsum()

ultima_emissao_lb_acum = df_curva["Emissão prevista na LB acumulada"].iloc[-1] if len(df_curva) > 0 else 1.0
if ultima_emissao_lb_acum == 0: ultima_emissao_lb_acum = 1.0

ultima_aprovacao_lb_acum = df_curva["Aprovação prevista na LB acumulada"].iloc[-1] if len(df_curva) > 0 else 1.0
if ultima_aprovacao_lb_acum == 0: ultima_aprovacao_lb_acum = 1.0

df_curva["Avanço da EI previsto na LB"] = df_curva["Emissão prevista na LB acumulada"] / ultima_emissao_lb_acum
df_curva["Avanço real da EI na LB"] = df_curva["Emissão realizada acumulada"] / ultima_emissao_lb_acum

diff_ei = df_curva["Avanço real da EI na LB"] - df_curva["Avanço da EI previsto na LB"]
df_curva["COMP_EI"] = np.where(diff_ei > 0, 1, np.where(diff_ei < 0, -1, 0))

df_curva["Avanço da AP previsto na LB"] = df_curva["Aprovação prevista na LB acumulada"] / ultima_aprovacao_lb_acum
df_curva["Avanço da AP real na LB"] = df_curva["Aprovação realizada acumulada"] / ultima_aprovacao_lb_acum

diff_ap = df_curva["Avanço da AP real na LB"] - df_curva["Avanço da AP previsto na LB"]
df_curva["COMP_AP"] = np.where(diff_ap > 0, 1, np.where(diff_ap < 0, -1, 0))

cols_curva_final = [
    "Data", "Emissão prevista na LB", "Emissão prevista na LB acumulada", "Emissão realizada por dia", "Emissão realizada acumulada",
    "Aprovação prevista na LB", "Aprovação realizada por dia", "Aprovação prevista na LB acumulada", "Aprovação realizada acumulada",
    "Avanço da EI previsto na LB", "Avanço real da EI na LB", "COMP_EI",
    "Avanço da AP previsto na LB", "Avanço da AP real na LB", "COMP_AP"
]
df_curva = df_curva[cols_curva_final]

# ============================================================
# 7. Salvar e Formatar
# ============================================================
data_str = datetime.datetime.now().strftime("%d-%m-%Y")
arquivo_saida = os.path.join(caminho_base, f"Comparativo Cronograma {data_str}.xlsx")
rev = 1
while os.path.exists(arquivo_saida):
    arquivo_saida = os.path.join(caminho_base, f"Comparativo Cronograma {data_str}_R{rev}.xlsx")
    rev += 1

with pd.ExcelWriter(arquivo_saida, engine="openpyxl") as writer:
    df_final.drop(columns=["pct_calc"]).to_excel(writer, sheet_name="Base", index=False)
    df_resumo_concl.to_excel(writer, sheet_name="Resumo", startrow=2, index=False)
    df_resumo_atras.to_excel(writer, sheet_name="Resumo", startrow=len(df_resumo_concl)+7, index=False)
    df_curva.to_excel(writer, sheet_name="Curva S", index=False)
    df_concluidas_list.drop(columns=["pct_calc"]).to_excel(writer, sheet_name="Concluídas", index=False)
    df_atrasadas_list.drop(columns=["pct_calc"]).to_excel(writer, sheet_name="Atrasadas", index=False)
    df_sprint_atual.to_excel(writer, sheet_name="Sprint Atual", index=False)
    df_prox_sprint.to_excel(writer, sheet_name="Próximo Sprint", index=False)

wb = load_workbook(arquivo_saida)

ws_res = wb["Resumo"]
ws_res["A1"] = "Número de documentos listados"
ws_res["B1"] = num_documentos
ws_res["A2"] = "Aderência do modelo (%)"
ws_res["B2"] = rbln_pct = data_ref # Preservado link lógico original
ws_res["B2"] = aderencia_pct
ws_res["B2"].number_format = '0.0%'

# Formatação de célula percentual
for row in range(4, len(df_resumo_concl) + 4):
    ws_res.cell(row=row, column=4).number_format = '0.0%' # Coluna D (%EI)
    ws_res.cell(row=row, column=8).number_format = '0.0%' # Coluna H (%AP)

# ============================================================
# 8. Gráficos Dinâmicos
# ============================================================
def add_stacked_chart(ws, start_row, end_row, title, pos, min_c=2, max_c=5):
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = title
    
    # Coleta de dados estruturada para ignorar métricas de proporção/mistas se necessário
    data = Reference(ws, min_col=min_c, max_col=max_c, min_row=start_row, max_row=end_row)
    cats = Reference(ws, min_col=1, min_row=start_row+1, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, pos)

# Ajuste analítico: Seleciona as colunas B e C (Totais de Emissão), pula a D (%EI) 
# e plota as colunas E, F, G (Avaliação, Atendimento e Aprovação) para evitar distorção de escala.
add_stacked_chart(ws_res, 3, len(df_resumo_concl)+3, "Concluídas por Disciplina", "J3", min_c=2, max_c=3)
add_stacked_chart(ws_res, len(df_resumo_concl)+8, len(df_resumo_concl)+len(df_resumo_atras)+8, "Atrasadas por Disciplina", "J20", min_c=2, max_c=5)

# Gráfico da Curva S
ws_curva = wb["Curva S"]
c_bar = BarChart()
c_bar.title = "Acompanhamento Emissão"
c_bar.y_axis.title = "Quantidade"
c_bar.x_axis = DateAxis(crossAx=100)
c_bar.x_axis.number_format = 'dd/mm'
c_bar.x_axis.majorTimeUnit = "days"

data_b = Reference(ws_curva, min_col=2, min_row=1, max_row=len(df_curva)+1)
data_d = Reference(ws_curva, min_col=4, min_row=1, max_row=len(df_curva)+1)
c_bar.add_data(data_b, titles_from_data=True)
c_bar.add_data(data_d, titles_from_data=True)

c_line = LineChart()
data_c = Reference(ws_curva, min_col=3, min_row=1, max_row=len(df_curva)+1)
data_e = Reference(ws_curva, min_col=5, min_row=1, max_row=len(df_curva)+1)
c_line.add_data(data_c, titles_from_data=True)
c_line.add_data(data_e, titles_from_data=True)
for s in c_line.series: 
    s.graphicalProperties.line.width = 30000 

c_bar += c_line
ws_curva.add_chart(c_bar, "G2")

wb.save(arquivo_saida)
print(f"\nProcessamento concluído com sucesso!")